import openpyxl
import subprocess
import shlex
from tqdm import tqdm
import atexit
import chardet

# Ścieżka do pliku XLSX z listą obiektów
xlsx_list_path = r'C:\Users\asabat\PycharmProjects\Searcher2\GCMObjects.xlsx'

# Ścieżka do katalogu, w którym chcesz przeszukać pliki
search_directory = r'C:\01_SRC\PORTAL'

# Lista katalogów, które chcesz wykluczyć z wyszukiwania
exclude_paths = [
 #   r'C:\4_2_X\apps\gcm\src\aix\dbdes',
    r'C:\inne_wykluczony_katalog',
    # Dodaj więcej ścieżek według potrzeb
]

# Ścieżka do pliku wynikowego
output_file_path = 'GCMObjects_in_PORTAL.xlsx'
output_count_file_path = 'GCMObjects_in_PORTAL_count.xlsx'  # Nowy plik na zliczenia

# Zmienna na listę pominiętych plików
skipped_files_list = []


# Funkcja do wczytywania fraz z pliku XLSX
def read_objects_from_xlsx(xlsx_file):
    objects = {}
    try:
        workbook = openpyxl.load_workbook(xlsx_file)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            object_name, object_type = row[0], row[1]
            if object_name and object_type:
                if object_type not in objects:
                    objects[object_type] = []
                objects[object_type].append(object_name)
    except Exception as e:
        print(f"Błąd wczytywania pliku XLSX: {e}")
    return objects


# Funkcja do przeszukiwania plików za pomocą ripgrep
def search_files(objects, directory, exclude_paths, progress_bar):
    results = {}
    count_results = {}  # Nowy słownik na zliczenia

    for object_type, object_names in objects.items():
        results[object_type] = []
        count_results[object_type] = {}  # Inicjalizacja słownika zliczeń

        for object_name in object_names:
            print(f"Wyszukiwanie frazy: {object_name}")

            # Exclude files with specific extensions (.dll, .svg) and additional paths
            exclude_path_args = ' '.join([f'--glob "!{path}/**"' for path in exclude_paths])
            cmd = f'rg {shlex.quote(object_name)} {directory} --no-heading --with-filename --line-number --glob "!*.dll" --glob "!*.svg" {exclude_path_args}'

            try:
                process = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                stdout, stderr = process.communicate()

                # Use chardet to detect encoding
                encoding = chardet.detect(stdout)['encoding']
                if encoding is None:
                    encoding = 'utf-8'  # Default to utf-8 if encoding detection fails

                # Decode stdout using detected encoding
                stdout = stdout.decode(encoding, errors='replace')

                if stderr:
                    print(f"Error while searching: {stderr.decode('utf-8')}")
                    continue  # Skip to the next iteration if an error occurs

                # Filtracja wyników przed wyświetleniem
                filtered_lines = [line for line in stdout.split('\n') if
                                  all(path not in line for path in exclude_paths)]

                # Zapisywanie wyników
                for line in filtered_lines:
                    if line:
                        try:
                            # Parse line number, file path, and line content
                            line_parts = line.split(':', 3)
                            line_number = line_parts[2]
                            file_path = line_parts[1].strip()
                            # Append only object name, file path, and line number
                            results[object_type].append((object_name, file_path, line_number))

                            # Zliczanie wyników
                            if object_name not in count_results[object_type]:
                                count_results[object_type][object_name] = 0
                            count_results[object_type][object_name] += 1

                        except Exception as e:
                            print(f"Error processing line: {e}")
                            continue  # Skip to the next iteration if an error occurs

                progress_bar.update(1)  # Aktualizacja paska postępu po zakończeniu przeszukiwania jednej frazy

            except Exception as e:
                print(f"Error while running command: {e}")
                continue

    return results, count_results


# Funkcja do zapisywania wyników do pliku XLSX
def save_results_to_excel(results, output_file):
    output_workbook = openpyxl.Workbook()

    for object_type, result_list in results.items():
        output_sheet = output_workbook.create_sheet(title=object_type)
        output_sheet.append(['Object Name', 'File Path', 'Line Number'])

        for result in result_list:
            output_sheet.append(result)

    # Usunięcie domyślnego arkusza
    del output_workbook['Sheet']

    output_workbook.save(output_file)


# Funkcja do zapisywania wyników zliczeń do pliku XLSX
def save_count_results_to_excel(count_results, objects, output_count_file):
    output_workbook = openpyxl.Workbook()

    for object_type, object_names in objects.items():
        output_sheet = output_workbook.create_sheet(title=object_type)
        output_sheet.append(['Object Name', 'Count'])

        for object_name in object_names:
            count = count_results[object_type].get(object_name, 0)
            output_sheet.append([object_name, count])

        # Dodanie fraz z 0 wystąpieniami, jeśli nie znaleziono ich w count_results
        for object_name in object_names:
            if object_name not in count_results[object_type] and object_name not in [row[0] for row in output_sheet.iter_rows(values_only=True)]:
                output_sheet.append([object_name, 0])

    # Usunięcie domyślnego arkusza
    del output_workbook['Sheet']

    output_workbook.save(output_count_file)


# Funkcja do zapisywania wyników przy przerwaniu programu
def save_results_on_exit():
    global search_results_dict, count_results_dict, skipped_files_list, output_file_path, output_count_file_path
    if search_results_dict:
        save_results_to_excel(search_results_dict, output_file_path)
    if count_results_dict:
        save_count_results_to_excel(count_results_dict, objects_dict, output_count_file_path)
    if skipped_files_list:
        save_skipped_files_to_excel(skipped_files_list, output_file_path)


# Wczytywanie obiektów z pliku XLSX
try:
    objects_dict = read_objects_from_xlsx(xlsx_list_path)

    # Rejestracja funkcji do zapisywania wyników przy przerwaniu programu
    atexit.register(save_results_on_exit)

    # Pasek postępu
    total_phrases = sum(len(names) for names in objects_dict.values())
    progress_bar = tqdm(total=total_phrases, desc="Przeszukiwanie plików", unit="phrase")

    # Przeszukiwanie plików za pomocą ripgrep
    search_results_dict, count_results_dict = search_files(objects_dict, search_directory, exclude_paths, progress_bar)

    # Zamknięcie paska postępu
    progress_bar.close()

except FileNotFoundError:
    print(f"Plik XLSX z listą obiektów nie został znaleziony.")
except Exception as e:
    print(f"Wystąpił nieoczekiwany błąd: {e}")
